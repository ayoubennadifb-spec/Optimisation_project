"""Lecture rigoureuse de ``Donnees_MaghrebSteel.xlsx`` (les 9 onglets).

Le fichier Excel est l'unique source de données numériques. Ce module le lit
avec openpyxl et renvoie un objet :class:`Data` aux structures propres et typées,
en validant au passage la cohérence (présence des grades, lignes, familles...).

On NE code en dur aucune valeur métier : tout vient de l'Excel. Les seuls
repères en dur sont des libellés d'en-tête utilisés pour repérer les blocs.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from . import config


# --------------------------------------------------------------------------
# Structures de données
# --------------------------------------------------------------------------
@dataclass(frozen=True)
class Order:
    """Une commande du carnet (onglet Commandes)."""

    id: str
    client: str
    famille: str
    grade: str
    ep: float          # épaisseur (mm)
    larg: int          # largeur (mm)
    ton: float         # tonnage demandé (T)
    prix: float        # prix de vente négocié (MAD/T)
    sem: int           # semaine d'échéance (1..4)
    prio: str          # Haute / Normale / Basse


@dataclass
class StockNiveau:
    """Stock initial avec bornes de sécurité (min) et physique (max)."""

    initial: float
    mini: float
    maxi: float


@dataclass
class Data:
    """Conteneur de toutes les données du problème."""

    orders: list[Order]
    cadence: dict[tuple[str, str], float]          # (ligne, famille) -> T/jour
    rend: dict[str, dict[str, float]]              # process -> {r, chute, decl, nc}
    cout: dict[str, list[float]]                   # clé coût -> 7 valeurs (tranches)
    prix_hrc: dict[tuple[str, int], float]         # (grade, largeur) -> MAD/T
    dispo_hrc: dict[str, float]                    # grade -> T disponibles sur l'horizon
    largeurs_hrc: list[int]
    stock_pk: dict[str, StockNiveau]               # grade -> stock PK
    stock_interprocess: dict[str, StockNiveau]     # point -> stock
    stock_pf: dict[str, StockNiveau]               # famille -> stock produit fini
    arret: dict[tuple[str, int], float]            # (ligne, semaine) -> jours d'arrêt
    param: dict[str, object]                       # paramètres bruts (onglet Parametres)

    # --- paramètres dérivés (lus dans `param`, avec valeurs de repli sûres) ---
    prix_chute: float = 1800.0
    coef_decl: float = 0.5
    coef_nc: float = 0.2
    prix_zinc: float = 18000.0
    cons_zinc_hdg: float = 0.025
    cons_zinc_ppgi: float = 0.025
    prix_peinture: float = 12000.0
    cons_peinture: float = 0.01
    penalite_retard: dict[str, float] = field(default_factory=dict)
    cout_stock_interprocess: float = 25.0
    cout_stock_pf: float = 40.0

    # --- index pratiques ---
    def orders_in_scope(self) -> list[tuple[int, Order]]:
        """(indice, commande) pour les familles modélisées (Quarto exclu)."""
        return [(i, o) for i, o in enumerate(self.orders)
                if o.famille in config.ROUTES]


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def _is_num(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _f(v, default=0.0) -> float:
    return float(v) if _is_num(v) else float(default)


# --------------------------------------------------------------------------
# Lecture des onglets
# --------------------------------------------------------------------------
def _load_commandes(ws) -> list[Order]:
    orders: list[Order] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = row[0]
        if not isinstance(cid, str) or not cid.startswith("CMD"):
            continue
        if not _is_num(row[6]):           # tonnage absent -> ligne de titre
            continue
        orders.append(Order(
            id=cid, client=row[1], famille=row[2], grade=row[3],
            ep=float(row[4]), larg=int(row[5]), ton=float(row[6]),
            prix=float(row[7]), sem=int(row[8]), prio=row[9],
        ))
    return orders


def _load_cadences(ws) -> dict[tuple[str, str], float]:
    cadence: dict[tuple[str, str], float] = {}
    fams = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "Ligne / Famille":
            fams = list(row[1:6])         # [HRC DEC, CRC, HDG, PPGI, BACR]
            continue
        if fams and row[0] in config.LINES:
            for j, fam in enumerate(fams):
                v = row[1 + j]
                if _is_num(v):
                    cadence[(row[0], fam)] = float(v)
    return cadence


def _load_rendements(ws) -> dict[str, dict[str, float]]:
    rend: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] in config.LINES:
            rend[row[0]] = dict(r=_f(row[1]), chute=_f(row[2]),
                                decl=_f(row[3]), nc=_f(row[4]))
    return rend


def _load_couts(ws) -> dict[str, list[float]]:
    cout: dict[str, list[float]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        key = row[0]
        if isinstance(key, str) and any(_is_num(v) for v in row[1:8]):
            cout[key] = [_f(v) for v in row[1:8]]
    return cout


def _load_prix_hrc(ws) -> tuple[dict, dict, list]:
    prix_hrc: dict[tuple[str, int], float] = {}
    dispo: dict[str, float] = {}
    largeurs: list[int] = []
    mode = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "Grade \\ Largeur":
            largeurs = [int(v) for v in row[1:8] if _is_num(v)]
            mode = "prix"
            continue
        if row[0] == "Grade":
            mode = "dispo"
            continue
        if mode == "prix" and row[0] in config.GRADES:
            for j, L in enumerate(largeurs):
                prix_hrc[(row[0], L)] = _f(row[1 + j])
        elif mode == "dispo" and row[0] in config.GRADES:
            dispo[row[0]] = _f(row[1])
    return prix_hrc, dispo, largeurs


def _load_stocks(ws) -> tuple[dict, dict, dict]:
    """Trois sous-tables : stocks PK (par grade), interprocess, produits finis."""
    stock_pk: dict[str, StockNiveau] = {}
    stock_inter: dict[str, StockNiveau] = {}
    stock_pf: dict[str, StockNiveau] = {}
    mode = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        label = row[0]
        if not isinstance(label, str):
            continue
        low = label.lower()
        if low.startswith("stocks pk"):
            mode = "pk"; continue
        if low.startswith("stocks interprocess"):
            mode = "inter"; continue
        if low.startswith("stocks produits finis"):
            mode = "pf"; continue
        if label in ("Grade", "Point de stockage", "Famille"):
            continue                       # ligne d'en-tête de sous-table
        if not (_is_num(row[1]) and _is_num(row[2]) and _is_num(row[3])):
            continue
        niveau = StockNiveau(initial=_f(row[1]), mini=_f(row[2]), maxi=_f(row[3]))
        if mode == "pk" and label in config.GRADES:
            stock_pk[label] = niveau
        elif mode == "inter":
            stock_inter[label] = niveau
        elif mode == "pf" and label in config.FAMILIES:
            stock_pf[label] = niveau
    return stock_pk, stock_inter, stock_pf


def _load_arrets(ws) -> dict[tuple[str, int], float]:
    arret: dict[tuple[str, int], float] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] in config.LINES:
            for t in config.WEEKS:
                arret[(row[0], t)] = _f(row[t])
    return arret


def _load_parametres(ws) -> dict[str, object]:
    param: dict[str, object] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if isinstance(row[0], str) and row[0].strip():
            param[row[0].strip()] = row[1]
    return param


# --------------------------------------------------------------------------
# Point d'entrée
# --------------------------------------------------------------------------
def load_data(path: str | Path | None = None) -> Data:
    """Charge toutes les données depuis l'Excel et renvoie un :class:`Data`."""
    path = Path(path) if path else config.DATA_FILE
    if not path.exists():
        raise FileNotFoundError(f"Fichier de données introuvable : {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    orders = _load_commandes(wb["Commandes"])
    cadence = _load_cadences(wb["Cadences"])
    rend = _load_rendements(wb["Rendements"])
    cout = _load_couts(wb["Couts_Variables"])
    prix_hrc, dispo_hrc, largeurs = _load_prix_hrc(wb["Prix_HRC"])
    stock_pk, stock_inter, stock_pf = _load_stocks(wb["Stocks_Initiaux"])
    arret = _load_arrets(wb["Arrets_Planifies"])
    param = _load_parametres(wb["Parametres"])

    def p(key, default):
        return _f(param.get(key), default)

    data = Data(
        orders=orders, cadence=cadence, rend=rend, cout=cout,
        prix_hrc=prix_hrc, dispo_hrc=dispo_hrc, largeurs_hrc=largeurs,
        stock_pk=stock_pk, stock_interprocess=stock_inter, stock_pf=stock_pf,
        arret=arret, param=param,
        prix_chute=p("Prix de valorisation des chutes", 1800.0),
        coef_decl=p("Coefficient déclassé/conforme", 0.5),
        coef_nc=p("Coefficient non-conforme/conforme", 0.2),
        prix_zinc=p("Prix zinc", 18000.0),
        cons_zinc_hdg=p("Consommation zinc HDG", 0.025),
        cons_zinc_ppgi=p("Consommation zinc PPGI", 0.025),
        prix_peinture=p("Prix peinture (PPGI)", 12000.0),
        cons_peinture=p("Consommation peinture PPGI", 0.01),
        penalite_retard={
            "Haute": p("Pénalité retard commande Haute", 500.0),
            "Normale": p("Pénalité retard commande Normale", 200.0),
            "Basse": p("Pénalité retard commande Basse", 0.0),
        },
        cout_stock_interprocess=p("Coût stockage interprocess", 25.0),
        cout_stock_pf=p("Coût stockage produit fini", 40.0),
    )
    _sanity_check(data)
    return data


def _sanity_check(data: Data) -> None:
    """Garde-fous : on échoue tôt et clairement si une donnée clé manque."""
    assert len(data.orders) >= 60, f"Carnet trop court : {len(data.orders)} commandes"
    for g in config.GRADES:
        assert g in data.dispo_hrc, f"Dispo HRC manquante pour le grade {g}"
    for ligne in config.LINES:
        assert ligne in data.rend, f"Rendement manquant pour la ligne {ligne}"
    for key in ("PK", "CRMA", "CRMB", "BAF", "SKP",
                "LGA-HDG", "LGA-PPGI", "LGA-BACR", "LGB-HDG", "LGB-BACR"):
        assert key in data.cout, f"Coût variable manquant pour {key}"
    for fam in config.FAMILIES:
        assert fam in data.stock_pf, f"Stock produit fini manquant pour {fam}"


if __name__ == "__main__":      # petit récapitulatif lisible
    d = load_data()
    print(f"Commandes              : {len(d.orders)}")
    print(f"Dont périmètre modélisé : {len(d.orders_in_scope())}")
    print(f"Cadences (couples)     : {len(d.cadence)}")
    print(f"Dispo HRC              : {d.dispo_hrc}")
    print(f"Stock PF (familles)    : {list(d.stock_pf)}")
    print(f"Pénalités retard       : {d.penalite_retard}")
